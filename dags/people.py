import csv
from airflow import DAG
from airflow.operators.python import PythonOperator
from datetime import datetime

import psycopg2
from airflow.providers.postgres.operators.postgres import PostgresOperator
from airflow.providers.postgres.hooks.postgres import PostgresHook
from count_years import years_until
import json
import pandas
from datetime import datetime
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

default_args = {
    'owner': 'Tetiana',
    'start_date':datetime(2023, 1, 1)
}


def _read_people():
    hook = PostgresHook(postgres_conn_id='postgres')
    with open('/opt/data/people-100.csv', "rt") as file_obj:
        reader_obj = csv.reader(file_obj, delimiter=',')
        next(reader_obj, None)  
        rows = []
        for row in reader_obj:
            date_str = row[7]  
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')          
            rows.append([
                int(row[0]), #Index
                row[1].strip(), #User Id
                row[2].strip(), #First Name
                row[3].strip(), #Last Name
                row[4].strip(), #Sex
                row[5].strip(), #Email
                row[6].strip(), #Phone
                date_obj,  #date_of_birth
                row[8].strip(), #job_title
            ])
        target_fields = ['index', 'User_Id', 'First_Name', 'Last_Name', 'Sex', 'Email', 'Phone', 'Date_of_birth', 'Job_title']
        hook.insert_rows(table='people', rows=rows,
                         target_fields=target_fields,
                         commit_every=1000, replace=True, replace_index="index")
def _save_people_to_excel(ti):
    hook = PostgresHook(postgres_conn_id="postgres")
   
    people_file = '/opt/data/people.xlsx'
    if not Path(people_file).is_file():        
        wb = openpyxl.Workbook()
        ws = wb.active
        wb.save(people_file)
    else: 
        wb = openpyxl.load_workbook(people_file)
    if 'People' in wb.sheetnames:
        ws = wb['People']
    else:
        ws = wb.active
        ws.title = 'People'


    df = hook.get_pandas_df(sql="SELECT * FROM (SELECT DISTINCT ON (user_id) * FROM people ) AS unique_people ORDER BY index")
    df['Age'] = df['date_of_birth'].apply(years_until)
    
    data = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(data, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(people_file)
        

with DAG(dag_id='people',
         default_args = default_args, 
         start_date=datetime(2023, 1, 1),
         schedule_interval='@daily', 
         catchup=False) as dag:
    create_people_table = PostgresOperator(
        task_id='create_people_table',
        postgres_conn_id='postgres',
        sql='''
        CREATE TABLE IF NOT EXISTS people (
            Index INT NOT NULL primary key,
            User_Id varchar(100),
            First_Name varchar(100),
            Last_Name varchar(100),
            Sex varchar(100),
            Email varchar(100),
            Phone varchar(100),
            Date_of_birth Date,
            Job_title varchar(100)
        );
    '''

)

    
    read_people = PythonOperator(
        task_id="read_people",
        python_callable=_read_people
    )
    save_people_to_excel = PythonOperator(
        task_id="save_people_to_excel",
        python_callable=_save_people_to_excel
    )
    
   

create_people_table >> read_people >> save_people_to_excel