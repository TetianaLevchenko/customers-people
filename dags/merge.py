from airflow import DAG
from airflow.providers.postgres.operators.postgres import PostgresOperator
from datetime import datetime
from airflow.operators.python import PythonOperator
from airflow.providers.postgres.hooks.postgres import PostgresHook
import openpyxl
from pathlib import Path
from openpyxl.utils.dataframe import dataframe_to_rows

default_args = {
    'start_date': datetime(2023, 1, 1)
}

def _save_merge_to_excel():
    hook = PostgresHook(postgres_conn_id="postgres")
    merge_file = '/opt/data/merge_tables.xlsx'
   
    if not Path(merge_file).is_file():        
        wb = openpyxl.Workbook()
        ws = wb.active
        wb.save(merge_file)
    else:
        wb = openpyxl.load_workbook(merge_file)

    if 'Merge_table' in wb.sheetnames:
        ws = wb['Merge_table']
    else:
        ws = wb.active
        ws.title = 'Merge_table'

    df = hook.get_pandas_df(sql="SELECT * FROM merged_table")

     
    data = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(data, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(merge_file)
    
    
with DAG(dag_id='merge_table',
         start_date=datetime(2023, 1, 1),
         default_args = default_args,
         schedule_interval='@daily', 
         catchup=False) as dag:
    
    create_merge_table = PostgresOperator(
        task_id='create_merge_table',
        postgres_conn_id='postgres',
        sql="""
        CREATE TABLE IF NOT EXISTS merged_table AS
            SELECT
            people.user_id AS user_id,
            people.first_name AS first_name,
            people.last_name AS last_name,
            people.sex AS sex,
            people.email AS email,
            people.phone AS phone,
            people.date_of_birth AS date_of_birth,
            people.job_title AS job_title
            FROM people
            INNER JOIN customers ON people.user_id = customers.customer_id;
        
        """)
save_merge_to_excel = PythonOperator(
        task_id="save_customers_to_excel",
        python_callable=_save_merge_to_excel
    )
create_merge_table >> save_merge_to_excel

