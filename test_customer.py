import csv
from airflow import DAG
from airflow.operators.python import PythonOperator
from datetime import datetime
from airflow.providers.postgres.operators.postgres import PostgresOperator
from airflow.providers.postgres.hooks.postgres import PostgresHook
from country_code import fix_phone
import pandas
from datetime import datetime
import numpy as np
from pathlib import Path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

default_args = {
    'owner': 'Tetiana',
    'start_date':datetime(2023, 1, 1)
}

def _read_customers():
    hook = PostgresHook(postgres_conn_id='postgres')
    
    with open('/opt/data/customers-100.csv', "rt") as file_obj:
        reader_obj = csv.reader(file_obj, delimiter=',')
        
        next(reader_obj, None)  

        rows = []

        for row in reader_obj:
            date_str = row[10]  
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            phone_fixed = fix_phone(row[7], row[6])
            
            rows.append([
                int(row[0]), #index
                row[1].strip(),  #Customer Id
                row[2].strip(), #First Name
                row[3].strip(), #Last Name
                row[4].strip(), #Company
                row[5].strip(), #City
                row[6].strip(), #Country
                phone_fixed,
                row[9].strip(), #Email
                date_obj,
                row[11].strip()
            ])

            
        print(rows)

        hook.insert_rows(table='customers_test', rows=rows,
                        target_fields=['Index', 'Customer_Id', 'First_Name', 'Last_Name','Company', 'City', 'Country', 'Phone_1','Email', 'Subscription', 'Website'],
                        commit_every=1000, replace=True, replace_index="index")
        
        
def _save_customers_to_excel():
    hook = PostgresHook(postgres_conn_id="postgres")
    customers_file = '/opt/data/customers_test.xlsx'
   
    if not Path(customers_file).is_file():        
        wb = openpyxl.Workbook()
        ws = wb.active
        wb.save(customers_file)
    else:
        wb = openpyxl.load_workbook(customers_file)

    if 'Customers' in wb.sheetnames:
        ws = wb['Customers']
    else:
        ws = wb.active
        ws.title = 'Customers'

    df = hook.get_pandas_df(sql="SELECT * FROM (SELECT DISTINCT ON (customer_id) * FROM customers_test ) AS unique_customers ORDER BY index")
   
    
    
    data = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(data, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(customers_file)


        
with DAG(dag_id='test_customers',
         start_date=datetime(2023, 1, 1),
         default_args = default_args,
         schedule_interval='@daily', 
         catchup=False) as dag:
    
    create_customers_table = PostgresOperator(
        task_id='create_customers_table',
        postgres_conn_id='postgres',
        sql='''
        CREATE TABLE IF NOT EXISTS customers_test (
            Index INT NOT NULL primary key,
            Customer_Id varchar(100),
            First_Name varchar(100),
            Last_Name varchar(100),
            Company varchar(100),
            City varchar(100),
            Country varchar(100),
            Phone_1 varchar(100),
            Email varchar(100),
            Subscription Date,
            Website varchar(100)
        );
    '''

)

    read_customers = PythonOperator(
        task_id="read_customers",
        python_callable=_read_customers
    )
    
    save_customers_to_excel = PythonOperator(
        task_id="save_customers_to_excel",
        python_callable=_save_customers_to_excel
    )

create_customers_table >> read_customers >> save_customers_to_excel